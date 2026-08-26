import os
import xlrd
import openpyxl
import database
import tempfile
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, render_template, request, send_file, flash, redirect, jsonify

app = Flask(__name__)
app.secret_key = 'super_secret_key'
UPLOAD_FOLDER = tempfile.gettempdir()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def process_invoice_files(invoice_paths, output_path):
    # Cache workbooks in memory to avoid reading from disk multiple times
    workbooks = []
    
    for invoice_path in invoice_paths:
        try:
            wb_in = xlrd.open_workbook(invoice_path, on_demand=True)
            workbooks.append(wb_in)
        except Exception as e:
            continue

    styles_data = []
    style_db_cache = database.get_style_weights()

    # Pass 1: Pre-scan all uploaded invoices to find and cache any available style weights
    # This ensures that if a style is missing its weight in Invoice A but has it in Invoice B,
    # processing Invoice A will still succeed using Invoice B's weight.
    for wb_in in workbooks:
        for si in range(wb_in.nsheets):
            ws_in = wb_in.sheet_by_index(si)
            if ws_in.name == 'PL': continue
            for i in range(ws_in.nrows):
                try:
                    r = [ws_in.cell_value(i,j) for j in range(ws_in.ncols)]
                    original_style = str(r[1]).strip()
                    try:
                        net_wt = float(r[9])
                    except:
                        net_wt = 0.0
                    
                    if original_style and net_wt > 0:
                        style_db_cache[original_style] = net_wt
                except: pass

    # Pass 2: Actual extraction and calculation
    for wb_in in workbooks:
        # 1. Try to extract exact PL Net Weight total for proportional scaling
        wb_pl_net = 0.0
        if 'PL' in wb_in.sheet_names():
            ws_pl = wb_in.sheet_by_name('PL')
            pl_found = False
            for i in range(max(0, ws_pl.nrows - 50), ws_pl.nrows):
                for j in range(ws_pl.ncols):
                    val = str(ws_pl.cell_value(i, j)).strip().upper()
                    if 'NET WEIGHT' in val or 'N.W' in val:
                        for k in range(j+1, ws_pl.ncols):
                            try:
                                num = float(ws_pl.cell_value(i, k))
                                if num > 0:
                                    wb_pl_net += num
                                    pl_found = True
                                    break
                            except: pass
                        if pl_found: break
                if pl_found: break
                
        # 2. Extract styles from INV sheets
        inv_no = 'Unknown Invoice'
        buyer = 'Unknown Buyer'
        current_quality = "(1) 100% Organic Cotton (RM0104) (40s VL, / INTERLOCK )"
        
        wb_styles = []
        for si in range(wb_in.nsheets):
            ws_in = wb_in.sheet_by_index(si)
            if ws_in.name.startswith('INV'):
                for i in range(min(15, ws_in.nrows)):
                    r = [str(ws_in.cell_value(i,j)).strip() for j in range(ws_in.ncols)]
                    full = ' '.join(r)
                    if 'SKDT/' in full and inv_no == 'Unknown Invoice':
                        for v in r:
                            if 'SKDT/' in v and '/DT:' in v:
                                inv_no = v.split('/DT:')[0].strip()
                    if r[0].startswith('M/S.') and 'SREE KANAGA' not in r[0] and buyer == 'Unknown Buyer':
                        buyer = r[0]
        
        for si in range(wb_in.nsheets):
            ws_in = wb_in.sheet_by_index(si)
            if ws_in.name == 'PL': continue
            
            for i in range(ws_in.nrows):
                col0 = str(ws_in.cell_value(i, 0)).strip()
                if 'ORGANIC COTTON' in col0.upper():
                    current_quality = col0
                    
                r = [ws_in.cell_value(i,j) for j in range(ws_in.ncols)]
                try:
                    original_style = str(r[1]).strip()
                    try:
                        qty = float(r[5])
                    except:
                        qty = 0.0
                    try:
                        net_wt = float(r[9])
                    except:
                        net_wt = 0.0
                    
                    if original_style and qty > 0:
                        if net_wt > 0:
                            style_db_cache[original_style] = net_wt
                        else:
                            net_wt = style_db_cache.get(original_style, 0.0)
                        
                        if net_wt > 0:
                            wb_styles.append({
                                'style': original_style,
                                'qty': qty,
                                'net_wt': net_wt,
                                'inv_no': inv_no,
                                'buyer': buyer,
                                'quality': current_quality
                            })
                except:
                    pass
        
        # Calculate per-workbook ratio
        wb_inv_net = sum(d['qty'] * d['net_wt'] for d in wb_styles)
        wb_ratio = wb_pl_net / wb_inv_net if wb_pl_net > 0 and wb_inv_net > 0 else 1.0

        for style_d in wb_styles:
            style_d['ratio'] = wb_ratio
            styles_data.append(style_d)

    database.save_style_weights(style_db_cache)
    idfl_stock = database.get_idfl_stock()
    
    # Ensure original_weight is populated for the current session to avoid dynamic fallbacks
    for stock in idfl_stock:
        if 'original_weight' not in stock:
            stock['original_weight'] = stock['remaining_weight']
    
    def find_matching_stock(quality, required_weight):
        # Determine target product string and sheet based on quality
        q = quality.upper()
        
        target_prod = None
        target_sheet = None
        
        if 'WOVEN' in q and '100%' in q:
            target_prod = 'Woven Fabrics' # Partial match
            target_sheet = 'IDFL'
        elif '95%' in q and '5%' in q:
            target_prod = '95%'
            target_sheet = None # Search all, or NON-IDFL
        elif '100% ORGANIC COTTON' in q:
            target_prod = '100% Organic Cotton'
            target_sheet = 'NON-IDFL'
            
        # Sort stock by remaining_weight (lowest first) so smaller balances get consumed first
        sorted_stock = sorted(idfl_stock, key=lambda x: x.get('remaining_weight', 0))
        
        for s in sorted_stock:
            # Skip if exhausted or not enough weight
            if s.get('status') == 'Exhausted' or s.get('remaining_weight', 0) < required_weight:
                continue
                
            # Sheet filter
            if target_sheet and s.get('sheet') != target_sheet:
                continue
                
            prod_str = s.get('products', '').upper()
            if target_prod and target_prod.upper() in prod_str:
                return s
                
            # If no target prod identified but it's 100% organic cotton
            if not target_prod and '100% ORGANIC COTTON' in prod_str:
                return s
                
        return None

    wb_out = openpyxl.Workbook()

    ws_out = wb_out.active
    ws_out.title = 'Mass Balance Sheet'
    
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    ws_out.merge_cells('A1:W1')
    c = ws_out['A1']
    c.value = "Mass Balance Sheet"
    c.font = Font(bold=True, size=14)
    c.alignment = center_align
    
    ws_out.merge_cells('B3:I3'); ws_out['B3'] = "Production Capacity :"
    ws_out.merge_cells('L3:S3'); ws_out['L3'] = "Storage Capacity :"
    
    ws_out.merge_cells('B4:I4'); ws_out['B4'] = "Inword Data"
    ws_out.merge_cells('J4:U4'); ws_out['J4'] = "Outward Data"
    ws_out.merge_cells('V4:W4'); ws_out['V4'] = "Stock Details"
    
    headers = [
        "Sr.No ", "Suppliers Name ", "Product Name and Quality", "TC No(IDFL or Other CB)", 
        "Certified Weight(Kg)", "Net Wt (Kg.)", "Gross Weight (Kg.)", "Lot No/Batch No", 
        "Open Stock in Kgs.", "Raw Material used in Kg", "Product Name", "Loss(%)", 
        "Buyers Name", "Invoice No.", "Certified Weight(Kg)", "Net Wt(Kg)", "Gross Weight(Kg)", 
        "Supplementary Wt (Kg)", "Transport Details(BL No/Challan No)", "Standard", 
        "IDFL TC No.", "Raw Material (Kg)", " Finished Product (kg)"
    ]
    
    ws_out.append(headers)
    for col_idx in range(1, len(headers)+1):
        cell = ws_out.cell(row=5, column=col_idx)
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
        if 2 <= col_idx <= 9: cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        elif 10 <= col_idx <= 21: cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        else: cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    for i in range(1, len(headers)+1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
    ws_out.column_dimensions['B'].width = 30
    ws_out.column_dimensions['C'].width = 40
    ws_out.column_dimensions['J'].width = 30
    ws_out.column_dimensions['K'].width = 35

    row_num = 6
    
    # Sort by invoice number to keep items from the same invoice grouped together
    styles_data.sort(key=lambda x: str(x.get('inv_no', '')))
    
    for idx, data in enumerate(styles_data, 1):
        style = data['style']
        qty = data['qty']
        buyer = data['buyer']
        inv_no = data['inv_no']
        quality = data.get('quality', "(1) 100% Organic Cotton (RM0104) (40s VL, / INTERLOCK )")
        ratio = data.get('ratio', 1.0)
        
        single_piece_wt = data.get('net_wt')
        
        if single_piece_wt is not None:
            # Scale proportionally so the sum of fin_val matches the PL sheet total exactly
            finished_prod = single_piece_wt * qty * ratio
            
            # Loss percentage increased to 21%
            loss_pct = 0.21
            raw_used = finished_prod * (1 + loss_pct)
            
            supp_wt = finished_prod * 0.10
            cert_wt = finished_prod - supp_wt
            
            raw_val = round(raw_used, 3)
            cert_val = round(cert_wt, 3)
            supp_val = round(supp_wt, 3)
            fin_val = round(finished_prod, 3)
            
            # FIFO Stock Logic
            matched_stock = find_matching_stock(quality, raw_val)
            
            if matched_stock:
                # Use finished_prod for net_val (output weights), tc_number for IDFL TC No.
                net_val = fin_val
                tc_number = matched_stock['tc_number']
                original_weight = matched_stock['original_weight']
                
                # Deduct weight
                matched_stock['remaining_weight'] -= raw_val
            else:
                net_val = fin_val
                tc_number = ""
                original_weight = ""
                
        else:
            raw_val = cert_val = net_val = supp_val = fin_val = ""
            tc_number = ""
            original_weight = ""
            
        # Force buyer to be exactly M/S. AB DUNS for all invoices
        buyer = "M/S. AB DUNS"
        
        standard = "GOTS"
        
        row_values = [
            idx, # 1: Sr.No
            "Sri Shanmugavel Mills Private Limited Knitting Division", # 2: Suppliers Name
            quality, # 3: Product Name and Quality
            tc_number, # 4: TC No(IDFL or Other CB)
            original_weight, # 5: Certified Weight
            original_weight, # 6: Net Wt
            original_weight, # 7: Gross Weight
            "", # 8: Lot No
            "", # 9: Open Stock
            raw_val, # 10: Raw Material used in Kg
            style, # 11: Product Name (Style)
            "21.000%", # 12: Loss(%)
            buyer, # 13: Buyers Name
            inv_no, # 14: Invoice No.
            cert_val, # 15: Certified Weight
            net_val, # 16: Net Wt
            net_val, # 17: Gross Weight
            supp_val, # 18: Supplementary Wt
            "", # 19: Transport Details
            standard, # 20: Standard
            "", # 21: IDFL TC No.
            "", # 22: Raw Material
            fin_val # 23: Finished Product
        ]
        
        for col_idx, val in enumerate(row_values, 1):

            cell = ws_out.cell(row=row_num, column=col_idx)
            cell.value = val
            cell.border = border
            if type(val) in [int, float]:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                
        row_num += 1

    database.save_idfl_stock(idfl_stock)
    wb_out.save(output_path)

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'invoice_file' not in request.files:
        flash('No file uploaded.')
        return redirect(request.url)
    
    files = request.files.getlist('invoice_file')
    if not files or files[0].filename == '':
        flash('No files selected.')
        return redirect(request.url)
        
    filepaths = []
    for file in files:
        if file and file.filename.endswith('.xls'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)
            filepaths.append(filepath)
            
    if not filepaths:
        flash('Please upload valid .xls invoice files.')
        return redirect('/')
        
    output_filename = "MASS_BALANCE_COMBINED.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    
    try:
        process_invoice_files(filepaths, output_path)
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        flash(f"Error processing files: {str(e)}")
        return redirect('/')


@app.route('/api/idfl-stock', methods=['GET', 'POST'])
def handle_idfl_stock():
    if request.method == 'POST':
        data = request.json
        database.save_idfl_stock(data)
        return jsonify({'status': 'success'})
    return jsonify(database.get_idfl_stock())


@app.route('/api/style-weights', methods=['GET'])
def api_style_weights():
    weights = database.get_style_weights()
    # Format as list of objects for frontend table
    data = [{'style': k, 'weight': v} for k, v in weights.items()]
    # Sort alphabetically by style name
    data.sort(key=lambda x: x['style'])
    return jsonify(data)

@app.route('/api/upload-idfl', methods=['POST'])
def upload_idfl():
    if 'idfl_file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['idfl_file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    ext = '.xls' if file.filename.lower().endswith('.xls') else '.xlsx'
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'uploaded_idfl' + ext)
    file.save(filepath)
    
    # Parse it
    try:
        import xlrd, openpyxl
        stock = []
        
        if ext == '.xls':
            wb = xlrd.open_workbook(filepath)
            sheetnames = wb.sheet_names()
            if 'NON-IDFL' in sheetnames:
                ws = wb.sheet_by_name('NON-IDFL')
                for i in range(4, ws.nrows):
                    try:
                        r1 = str(ws.cell_value(i, 1)).strip()
                        r2 = ws.cell_value(i, 2)
                        r3 = str(ws.cell_value(i, 3)).strip()
                        if not r1 or not r3: continue
                        if r1.upper() == 'TC NUMBER': continue
                        try: rem = float(r2)
                        except: continue
                        stock.append({
                            'id': f'non_idfl_{i}',
                            'sheet': 'NON-IDFL',
                            'tc_number': r1,
                            'initial_weight': rem,
                            'remaining_weight': rem,
                            'original_weight': rem,
                            'products': r3,
                            'status': 'Active'
                        })
                    except: pass
            if 'IDFL' in sheetnames:
                ws = wb.sheet_by_name('IDFL')
                for i in range(6, ws.nrows):
                    try:
                        r2 = str(ws.cell_value(i, 2)).strip()
                        r3 = str(ws.cell_value(i, 3)).strip()
                        r5 = ws.cell_value(i, 5)
                        r6 = str(ws.cell_value(i, 6)).strip()
                        if not r2 or not r6: continue
                        if r2.upper() == 'TC NUMBER': continue
                        try: rem = float(r5)
                        except: continue
                        stock.append({
                            'id': f'idfl_{i}',
                            'sheet': 'IDFL',
                            'tc_number': r2,
                            'initial_weight': rem,
                            'remaining_weight': rem,
                            'original_weight': rem,
                            'products': r6,
                            'status': r3 if r3 else 'Active'
                        })
                    except: pass
        else:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if 'NON-IDFL' in wb.sheetnames:
                ws = wb['NON-IDFL']
                for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
                    if not row[1] or not row[3]: continue
                    tc = str(row[1]).strip()
                    if tc.upper() == 'TC NUMBER': continue
                    try: rem = float(row[2])
                    except: continue
                    prod = str(row[3]).strip()
                    stock.append({
                        'id': f'non_idfl_{i}',
                        'sheet': 'NON-IDFL',
                        'tc_number': tc,
                        'initial_weight': rem,
                        'remaining_weight': rem,
                        'original_weight': rem,
                        'products': prod,
                        'status': 'Active'
                    })
            if 'IDFL' in wb.sheetnames:
                ws = wb['IDFL']
                for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True)):
                    if not row[2] or not row[6]: continue
                    tc = str(row[2]).strip()
                    if tc.upper() == 'TC NUMBER': continue
                    try: rem = float(row[5])
                    except: continue
                    prod = str(row[6]).strip()
                    stock.append({
                        'id': f'idfl_{i}',
                        'sheet': 'IDFL',
                        'tc_number': tc,
                        'initial_weight': rem,
                        'remaining_weight': rem,
                        'original_weight': rem,
                        'products': prod,
                        'status': str(row[3]).strip() if row[3] else 'Active'
                    })
        
        database.save_idfl_stock(stock)
        return jsonify({'status': 'success', 'records': len(stock)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
